import kivy
kivy.require('1.10.1')

from kivy.app import App
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.properties import ListProperty, NumericProperty, StringProperty, DictProperty, ObjectProperty, BooleanProperty
from kivy.core.window import Window
from kivy.clock import Clock 

from kivy.uix.dropdown import DropDown
from kivy.uix.button import Button

import os
import time
import yaml
from datetime import datetime, date
from dateutil import relativedelta
from copy import deepcopy

from openpyxl import Workbook 
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle

from screeninfo import get_monitors

m = get_monitors()[0]
Window.size = (m.width * 0.9 , m.height * 0.9)
Window.left = m.width * 0.05
Window.top = m.height * 0.05

path = os.path.dirname(os.path.realpath(__file__))

info = {}
config = {}

bold = Font(bold=True, name='Calibri')
center = Alignment(horizontal='center', vertical='center')
date_style = NamedStyle(name='date', number_format='DD/MM/YYYY')

default_save_dir = os.path.join(os.path.expanduser('~'), 'Neuropsy')
if not os.path.exists(default_save_dir):
    os.makedirs(default_save_dir)

with open('main.kv', encoding='utf-8') as f:
    Builder.load_string(f.read())

sm = ScreenManager()
instructions_key = ''

def string_to_date(string):
    day, month, year = tuple([int(x) for x in string.replace(' ', '').split('/')])
    if year <= datetime.now().year % 100:
        year += 2000  
    elif year < 100:
        year += 1900

    return date(year, month, day)

def adjust_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

def save_data(log, update_sheet):
    dob = string_to_date(info['nasc'])
    difference = relativedelta.relativedelta(datetime.now(), dob)
    now = datetime.now()

    write_info = deepcopy(info)

    write_info['nasc'] = f'{dob.day}/{dob.month}/{dob.year}'
    write_info['idade_anos'] = difference.years
    write_info['idade_meses'] = difference.years * 12 + difference.months
    write_info['date'] = f'{now.day}/{now.month}/{now.year}'

    del write_info['save_dir']
    data = {
        'user': write_info,
        'log': log
    }

    save_path = os.path.join(path, info['save_dir'], f'{write_info["nome"]} {dob.day}-{dob.month}-{dob.year}.xlsx')
    if os.path.exists(save_path):
        wb = load_workbook(save_path)
    else:
        wb = Workbook()

    sheets = ['flanker', 'memoria']
    for name in wb.sheetnames:
        if name not in sheets:
            del wb[name]
    
    ws = wb.create_sheet(title='usuario')
    ws.dimensions

    ws['A1'] = 'Informacoes do Usuario'
    ws['A1'].font = bold
    ws['A1'].alignment = center
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.merge_cells('A1:B1')

    ws['A2'] = 'Nome'
    ws['B2'] = data['user']['nome']
    
    ws['A3'] = 'Data do Teste'
    ws['B3'] = data['user']['date']
    ws['B3'].number_format='DD/MM/YYYY'

    ws['A4'] = 'Data de Nascimento'
    ws['B4'] = data['user']['nasc']
    ws['B4'].number_format='DD/MM/YYYY'

    ws['A5'] = 'Regiao'
    ws['B5'] = data['user']['regiao']

    ws['A6'] = 'Idade (Anos)'
    ws['B6'] = data['user']['idade_anos']

    ws['A7'] = 'Idade (Meses)'
    ws['B7'] = data['user']['idade_meses']

    if update_sheet == 'flanker':
        if 'flanker' in wb.sheetnames:
            del wb['flanker']
        ws = wb.create_sheet(title='flanker')
        ws.page_setup.fitToWidth = 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10

        ws.column_dimensions['D'].width = 2
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        ws['A1'] = 'Resultados'
        ws['A1'].font = bold
        ws['A1'].alignment = center
        ws.merge_cells('A1:F1')
        ws['A2'] = 'Resposta do Usuario'
        ws['A2'].font = bold
        ws['A2'].alignment = center
        ws['B2'] = 'Resposta Desejada'
        ws['B2'].font = bold
        ws['B2'].alignment = center
        ws['C2'] = 'Tempo'
        ws['C2'].font = bold
        ws['C2'].alignment = center
        for idx, elem in enumerate(data['log']):
            ws.cell(column=1, row=idx + 3, value=elem['res_user'])
            ws.cell(column=2, row=idx + 3, value=elem['res_actual'])
            ws.cell(column=3, row=idx + 3, value=elem['time'])

        ws['E2'] = 'Resumo'
        ws.merge_cells('E2:F2')
        ws['E2'].font = bold
        ws['E2'].alignment = center
        ws['E3'] = 'Acertos'
        ws['E4'] = 'Tempo Total(s)'
        ws['E5'] = 'Tempo (média)'

        ws['F3'] = '=SUMPRODUCT((A3:A99=B3:B99)*(A3:A99<>""))'
        ws['F4'] = '=SUM(C3:C99)'
        ws['F5'] = '=F4/COUNT(C3:C99)'

    elif update_sheet=='memory':
        if 'memoria' in wb.sheetnames:
            del wb['memoria']
        ws = wb.create_sheet(title='memoria')

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10

        ws.column_dimensions['C'].width = 2
        ws.column_dimensions['D'].width = 2
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        ws['A1'] = 'Resultados'
        ws.merge_cells('A1:F1')
        ws['A1'].font = bold
        ws['A1'].alignment = center
        ws['A2'] = 'Resposta'
        ws['A2'].font = bold
        ws['A2'].alignment = center
        ws['B2'] = 'Tempo'
        ws['B2'].font = bold
        ws['B2'].alignment = center
        for idx, elem in enumerate(data['log']):
            ws.cell(column=1, row=idx + 3, value=elem['res'])
            ws.cell(column=2, row=idx + 3, value=elem['time'])

        ws['E2'] = 'Resumo'
        ws.merge_cells('E2:F2')
        ws['E2'].font = bold
        ws['E2'].alignment = center
        ws['E3'] = 'Acertos'
        ws['E4'] = 'Tempo Total (s)'
        ws['E5'] = 'Tempo (Media)'

        ws['F3'] = '=COUNTIF(A3:A99, "certo")'
        ws['F4'] = '=SUM(B3:B99)'
        ws['F5'] = '=F4/COUNT(B3:B99)'

    try:
        wb.save(filename=save_path)
        wb.close()
    except Exception as e:
        p, ext = os.path.splitext(save_path)
        for i in range(100):
            new_path = f'{p} ({i}){ext}'
            if not os.path.exists(new_path):
                wb.save(filename=new_path)
                wb.close()
                return 
        
    


def input_valid():
    try:
        if 'nome' not in info or not info['nome']: 
            raise Exception() 
        if 'regiao' not in info or not info['regiao']: 
            raise Exception() 
        dob = string_to_date(info['nasc'])
        if 'save_dir' not in info or not os.path.isdir(info['save_dir']):
            raise Exception()
    except Exception as e:
        return False
    return True 

class Start(Screen):
    save_dir = StringProperty(default_save_dir)
    games_disabled = BooleanProperty(False)
        
    def on_enter(self):
        global config
        config = {}
        self.update('save_dir', self.save_dir)

    def update(self, key, value):
        global info
        if key == 'save_dir':
            self.save_dir = value
        info[key] = value

        self.games_disabled = not input_valid()

    def start_game(self, game):
        global config, instructions_key
        config_path = os.path.join(path, 'data', f'{game}.yml')
        config = yaml.load(open(config_path, encoding='utf-8'))

        instructions_key = 'instructions'
        sm.current = 'instruction'

class Flanker(Screen):
    content = ListProperty()
    right_image = StringProperty()
    left_image = StringProperty()
    counter = NumericProperty()

    def _keyboard_closed(self):
        if(self._keyboard):
            self._keyboard.unbind(on_key_down=self._on_keyboard_down)
            self._keyboard = None

    def _on_keyboard_down(self, keyboard, keycode, text, modifiers):
        if keycode[1] in self.keys and len(self.log) < len(self.content):       
            direction = keycode[1] in self.right_keys
            dt = time.time() - self.start_time
            self.log.append((direction, dt))
        
            if self.counter + 1 >= len(self.content):
                sm.current = 'start'
                return True  
            self.counter += 1
            self.start_time = time.time()
        return True

    def on_pre_enter(self):
        self.left_keys = {'q', 'left'}
        self.right_keys = {'p', 'right'}
        self.keys = self.left_keys | self.right_keys
        self.left_image = os.path.join(path, 'data', config['left_image'])
        self.right_image = os.path.join(path, 'data', config['right_image'])
        self.middle = True
        self.crowd = True
        self.content = config['content']
        self.counter = 0

        self.start_time = time.time()
        self.log = []
        self._keyboard = Window.request_keyboard(self._keyboard_closed, self)
        self._keyboard.bind(on_key_down=self._on_keyboard_down)

    def on_leave(self):
        log_data = []
        for i in range(len(self.log)):
            user_dir, dt = self.log[i]
            ans_dir = self.content[i][1]
            log_data.append({
                'res_user': 'direita' if user_dir else 'esquerda',
                'res_actual': 'direita' if ans_dir else 'esquerda',
                'time': dt
            })

        save_data(log_data, 'flanker')

        if(self._keyboard):
            self._keyboard.unbind(on_key_down=self._on_keyboard_down)

class Memory(Screen):
    counter = NumericProperty()
    inner_counter = NumericProperty()
    content = ListProperty()
    order = ListProperty()
    paths = DictProperty()
   
    def update(self, dt):
        if self.inner_counter + 1 < len(self.content[self.counter]):
            self.inner_counter += 1
            if self.inner_counter + 1 >= len(self.content[self.counter]):
                self.interval.cancel()
                self.start_time = time.time()
    
    def _keyboard_closed(self):
        if(self._keyboard):
            self._keyboard.unbind(on_key_down=self._on_keyboard_down)
            self._keyboard = None

    def _on_keyboard_down(self, keyboard, keycode, text, modifiers):
        if (keycode[1] in self.keys 
            and self.inner_counter + 1 >= len(self.content[self.counter])
            and len(self.log) < len(self.content)):

            if self.counter + 1 >= len(self.content):
                global instructions_key
                if instructions_key == 'instructions':
                    self.inner_counter = 0
                    self.counter = 0
                    instructions_key = 'instructions2'
                    sm.current = 'instruction'
                else:  
                    log_data = []
                    log = [item for sublist in self.logs for item in sublist]
                    for i in range(len(log)):
                        is_true, dt = log[i]
                        log_data.append({
                            'res': 'certo' if is_true else 'errado',
                            'time': dt
                        })

                    save_data(log_data, 'memory')
                    sm.current = 'start'
                return True

            is_true = keycode[1] in self.true_keys
            dt = time.time() - self.start_time
            self.log.append((is_true, dt))

            self.inner_counter = 0
            self.counter += 1
            self.interval = Clock.schedule_interval(self.update, 2.0)
              
        return True
    
    def on_pre_enter(self):
        self.inner_counter = 0
        self.counter = 0

        if instructions_key == 'instructions':
            paths = {}
            for entry in config['order']:
                paths[entry['name']] = os.path.join(path, 'data', entry['path'])

            self.logs = []
            self.paths = paths 
            self.content = config['content']

            self.true_keys = {'1', 'numpad1', 'right'}
            self.false_keys = {'0', 'numpad0', 'left'}
            self.keys = self.true_keys | self.false_keys
        elif instructions_key == 'instructions2':
            self.content = config['content2']
        
        self.logs.append([])
        self.log = self.logs[-1]

        self.start_time = time.time()
        self._keyboard = Window.request_keyboard(self._keyboard_closed, self)
        self._keyboard.bind(on_key_down=self._on_keyboard_down)


    def on_enter(self):
        self.interval = Clock.schedule_interval(self.update, 2.0)
        
    def on_leave(self): 
        if(self._keyboard):
            self._keyboard.unbind(on_key_down=self._on_keyboard_down)
            self.interval.cancel()


class Instruction(Screen):
    instructions = ListProperty([])
    counter = NumericProperty(0)

    def _keyboard_closed(self):
        if(self._keyboard):
            self._keyboard.unbind(on_key_down=self._on_keyboard_down)
            self._keyboard = None

    def _on_keyboard_down(self, keyboard, keycode, text, modifiers):
        print(keycode[1])
        if (keycode[1] in ['spacebar', 'right']):
            self.next()
        elif (keycode[1] in ['left']):
            self.previous()
    
    def on_pre_enter(self):
        global info, config
        self.info = info 
        instructions = config[instructions_key]  
        for instruction in instructions:
            instruction['image'] = os.path.join(path, 'data', instruction['image'])
        
        self.counter = 0
        self.instructions = instructions
        
        self._keyboard = Window.request_keyboard(self._keyboard_closed, self)
        self._keyboard.bind(on_key_down=self._on_keyboard_down)

    def next(self):
        if self.counter + 1 >= len(self.instructions):
            sm.current = config['game']
        else:
            self.counter += 1

    def previous(self):
        if self.counter - 1 >= 0:
            self.counter -= 1

    def on_leave(self):
        if (self._keyboard):
            self._keyboard.unbind(on_key_down=self._on_keyboard_down)



sm.add_widget(Start(name='start'))
sm.add_widget(Flanker(name='flanker'))
sm.add_widget(Memory(name='memory'))
sm.add_widget(Instruction(name = 'instruction'))

class NeuropsyApp(App):
    icon = 'icon.png'

    def build(self):
        return sm

if __name__ == '__main__':
    NeuropsyApp().run()