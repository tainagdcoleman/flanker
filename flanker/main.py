import kivy
kivy.require('1.10.1')

from kivy.app import App
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.properties import ListProperty, NumericProperty, StringProperty, DictProperty, ObjectProperty, BooleanProperty
from kivy.core.window import Window
from kivy.clock import Clock 

from kivy.uix.dropdown import DropDown
from kivy.uix.button import Button

from kivy.config import Config
Config.set('kivy', 'exit_on_escape', 0)

import os
import sys
import time
import yaml
from datetime import datetime, date
from dateutil import relativedelta
from copy import deepcopy

from openpyxl import Workbook 
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle

from screeninfo import get_monitors

m = get_monitors()[0]
Window.size = (m.width, m.height)
Window.fullscreen = True

path = os.path.dirname(os.path.realpath(__file__))

info = {}
game = ''

bold = Font(bold=True, name='Calibri')
center = Alignment(horizontal='center', vertical='center')
date_style = NamedStyle(name='date', number_format='DD/MM/YYYY')

default_save_dir = os.path.join(os.path.expanduser('~'), 'Neuropsy')
if not os.path.exists(default_save_dir):
    os.makedirs(default_save_dir)

with open(os.path.join(path, 'main.kv'), encoding='utf-8') as f:
    Builder.load_string(f.read())

sm = ScreenManager()
instructions_key = ''

def string_to_date(string):
    day, month, year = tuple([int(x) for x in string.replace(' ', '').split('/')])
    if year <= datetime.now().year % 100:
        year += 2000  
    elif year < 100:
        year += 1900

    return date(year, month, day)

def adjust_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

def save_data(log):
    global game
    dob = string_to_date(info['nasc'])
    difference = relativedelta.relativedelta(datetime.now(), dob)
    now = datetime.now()

    write_info = deepcopy(info)

    write_info['nasc'] = f'{dob.day}/{dob.month}/{dob.year}'
    write_info['idade_anos'] = difference.years
    write_info['idade_meses'] = difference.years * 12 + difference.months
    write_info['date'] = f'{now.day}/{now.month}/{now.year}'

    del write_info['save_dir']
    data = {
        'user': write_info,
        'log': log
    }

    save_path = os.path.join(path, info['save_dir'], f'{write_info["nome"]} {dob.day}-{dob.month}-{dob.year}.xlsx')
    if os.path.exists(save_path):
        wb = load_workbook(save_path)
    else:
        wb = Workbook()

    sheets = ['flanker', 'memoria']
    for name in wb.sheetnames:
        if name not in sheets:
            del wb[name]
    
    ws = wb.create_sheet(title='usuario')
    ws.dimensions

    ws['A1'] = 'Informacoes do Usuario'
    ws['A1'].font = bold
    ws['A1'].alignment = center
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.merge_cells('A1:B1')

    ws['A2'] = 'Nome'
    ws['B2'] = data['user']['nome']
    
    ws['A3'] = 'Data do Teste'
    ws['B3'] = data['user']['date']
    ws['B3'].number_format='DD/MM/YYYY'

    ws['A4'] = 'Data de Nascimento'
    ws['B4'] = data['user']['nasc']
    ws['B4'].number_format='DD/MM/YYYY'

    ws['A5'] = 'Regiao'
    ws['B5'] = data['user']['regiao']

    ws['A6'] = 'Idade (Anos)'
    ws['B6'] = data['user']['idade_anos']

    ws['A7'] = 'Idade (Meses)'
    ws['B7'] = data['user']['idade_meses']

    if game == 'flanker':
        if 'flanker' in wb.sheetnames:
            del wb['flanker']
        ws = wb.create_sheet(title='flanker')
        ws.page_setup.fitToWidth = 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10

        ws.column_dimensions['D'].width = 2
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        ws['A1'] = 'Resultados'
        ws['A1'].font = bold
        ws['A1'].alignment = center
        ws.merge_cells('A1:F1')
        ws['A2'] = 'Resposta do Usuario'
        ws['A2'].font = bold
        ws['A2'].alignment = center
        ws['B2'] = 'Resposta Desejada'
        ws['B2'].font = bold
        ws['B2'].alignment = center
        ws['C2'] = 'Tempo'
        ws['C2'].font = bold
        ws['C2'].alignment = center
        for idx, elem in enumerate(data['log']):
            ws.cell(column=1, row=idx + 3, value=elem['res_user'])
            ws.cell(column=2, row=idx + 3, value=elem['res_actual'])
            ws.cell(column=3, row=idx + 3, value=elem['time'])

        ws['E2'] = 'Resumo'
        ws.merge_cells('E2:F2')
        ws['E2'].font = bold
        ws['E2'].alignment = center
        ws['E3'] = 'Acertos'
        ws['E4'] = 'Tempo Total(s)'
        ws['E5'] = 'Tempo (média)'

        ws['F3'] = '=SUMPRODUCT((A3:A99=B3:B99)*(A3:A99<>""))'
        ws['F4'] = '=SUM(C3:C99)'
        ws['F5'] = '=F4/COUNT(C3:C99)'

    elif game=='memory':
        if 'memoria' in wb.sheetnames:
            del wb['memoria']
        ws = wb.create_sheet(title='memoria')

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10

        ws.column_dimensions['C'].width = 2
        ws.column_dimensions['D'].width = 2
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        ws['A1'] = 'Resultados'
        ws.merge_cells('A1:F1')
        ws['A1'].font = bold
        ws['A1'].alignment = center
        ws['A2'] = 'Resposta'
        ws['A2'].font = bold
        ws['A2'].alignment = center
        ws['B2'] = 'Tempo'
        ws['B2'].font = bold
        ws['B2'].alignment = center
        for idx, elem in enumerate(data['log']):
            ws.cell(column=1, row=idx + 3, value=elem['res'])
            ws.cell(column=2, row=idx + 3, value=elem['time'])

        ws['E2'] = 'Resumo'
        ws.merge_cells('E2:F2')
        ws['E2'].font = bold
        ws['E2'].alignment = center
        ws['E3'] = 'Acertos'
        ws['E4'] = 'Tempo Total (s)'
        ws['E5'] = 'Tempo (Media)'

        ws['F3'] = '=COUNTIF(A3:A99, "certo")'
        ws['F4'] = '=SUM(B3:B99)'
        ws['F5'] = '=F4/COUNT(B3:B99)'

    try:
        wb.save(filename=save_path)
        wb.close()
    except Exception as e:
        p, ext = os.path.splitext(save_path)
        for i in range(100):
            new_path = f'{p} ({i}){ext}'
            if not os.path.exists(new_path):
                wb.save(filename=new_path)
                wb.close()
                return 
        
def input_valid():
    try:
        if 'nome' not in info or not info['nome']: 
            raise Exception() 
        if 'regiao' not in info or not info['regiao']: 
            raise Exception() 
        dob = string_to_date(info['nasc'])
        if 'save_dir' not in info or not os.path.isdir(info['save_dir']):
            raise Exception()
    except Exception as e:
        return False
    return True 

class Start(Screen):
    save_dir = StringProperty(default_save_dir)
    games_disabled = BooleanProperty(False)

    def on_enter(self):
        self.update('save_dir', self.save_dir)

    def update(self, key, value):
        global info
        if key == 'save_dir':
            self.save_dir = value
        info[key] = value

        self.games_disabled = not input_valid()

    def start_game(self, chosen_game):
        global game, instructions_key
        game = chosen_game

        instructions_key = 'instructions'
        sm.current = 'game'

    def exit_app(self):
        sys.exit(0)

class Game(Screen):
    image = StringProperty()

    def on_pre_enter(self):
        global game
        config_path = os.path.join(path, 'data', game, 'config.yml')
        self.config = yaml.load(open(config_path, encoding='utf-8'))
        
        self._keyboard = Window.request_keyboard(self._keyboard_closed, self)
        self._keyboard.bind(on_key_down=self._on_keyboard_down)

        if game == 'flanker':
            self.init_flanker()
        elif game == 'memory':
            self.init_memory()
        else:
            sm.current = 'start'
            
    def _on_keyboard_down(self, keyboard, keycode, text, modifiers):
        global game
        if keycode[1] == 'escape':
            sm.current = 'start'

        if game == 'flanker':
            self.keyboard_flanker(keycode[1])
        elif game == 'memory':
            self.keyboard_memory(keycode[1])
        else:
            sm.current = 'start'

        return True

    def _keyboard_closed(self):
        if(self._keyboard):
            self._keyboard.unbind(on_key_down=self._on_keyboard_down)
            self._keyboard = None

    def on_leave(self):
        global game
        if game == 'flanker':
            self.finish_flanker()
        elif game == 'memory':
            self.finish_memory()

        if(self._keyboard):
            self._keyboard.unbind(on_key_down=self._on_keyboard_down)

    def get_slide(self):
        global path
        return os.path.join(path, 'data', game, f'Slide{self.index + 1}.PNG')

    ### FLANKER GAME ###
    def init_flanker(self):
        self.index = -1
        self.logs = []
        self.left_keys = {'q', 'left'}
        self.right_keys = {'p', 'right'}
        self.keys = self.left_keys | self.right_keys
        self.next_flanker()

    def keyboard_flanker(self, keycode):
        slide = self.config['slides'][self.index]
        if keycode == 'spacebar' and slide['type'] == 'instruction':
            self.next_flanker()
        elif keycode in self.keys and slide['type'] == 'content':
            dt = time.time() - self.time
            user_answer = 'direita' if keycode in self.right_keys else 'esquerda'
            answer = self.config['slides'][self.index]['answer']
            answer = 'direita' if answer == 'right' else 'esquerda'
            self.logs.append((user_answer, answer, dt))
            self.next_flanker()

    def next_flanker(self):
        if self.index + 1 >= len(self.config['slides']):
            sm.current = 'start'
        else:
            self.index += 1

            slide = self.config['slides'][self.index]
            if slide['type'] == 'instruction':
                pass
            elif slide['type'] == 'content':
                self.time = time.time()
            self.image = self.get_slide()
    
    def finish_flanker(self):
        log_data = []
        for user_answer, answer, dt in self.logs:
            log_data.append({
                'res_user': user_answer,
                'res_actual': answer,
                'time': dt
            })

        save_data(log_data)

    ### MEMORY GAME ###        
    def init_memory(self):
        self.index = -1
        self.interval = None
        self.logs = []
        self.wrong_keys = {'1', 'numpad1'}
        self.right_keys = {'0', 'numpad0'}
        self.keys = self.wrong_keys | self.right_keys
        self.next_memory()

    def keyboard_memory(self, keycode):
        slide = self.config['slides'][self.index]
        if keycode == 'spacebar' and slide['type'] == 'instruction':
            self.next_memory()
        elif keycode in self.keys and slide['type'] == 'wait_for_feedback':
            dt = time.time() - self.time
            user_answer = 'certo' if keycode in self.right_keys else 'errado'
            self.logs.append({'res': user_answer, 'time': dt})
            self.next_memory()

    def next_memory(self, *args, **kwargs):
        if self.index + 1 >= len(self.config['slides']):
            sm.current = 'start'
        else:
            self.index += 1

            slide = self.config['slides'][self.index]
            if slide['type'] == 'instruction':
                pass
            if slide['type'] == 'wait_for_feedback':
                if self.interval is not None:
                    self.interval.cancel()
                    self.interval = None
                self.time = time.time()
            elif slide['type'] == 'content':
                if self.interval is None:
                    self.interval = Clock.schedule_interval(self.next_memory, 2.0)
            self.image = self.get_slide()
    
    def finish_memory(self):
        save_data(self.logs)

sm.add_widget(Start(name='start'))
sm.add_widget(Game(name='game'))

class NeuropsyApp(App):
    icon = 'icon.png'

    def build(self):
        return sm

if __name__ == '__main__':
    NeuropsyApp().run()